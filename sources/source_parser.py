import logging

import openpyxl
from openpyxl.workbook import Workbook

FILE_PATH: str = "Таблица с дублями кейсов.xlsx"
CASES_BORDER: int = 6234

logger = logging.getLogger(__name__)


def get_sources(file_path: str, cases_border: int):
    logger.info('Processing xlsx source file...')
    workbook: Workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    max_row: int = sheet.max_row
    old_cases_ids: list[str] = []
    new_cases_ids: list[str] = []
    for i in range(2, max_row):
        if i <= cases_border - 1:
            old_cases_ids.append(sheet.cell(i, 2).value)
        else:
            new_cases_ids.append(sheet.cell(i, 2).value)

    logger.info('Xlsx source file processed!')
    return old_cases_ids, new_cases_ids


if __name__ == '__main__':
    old_cases, new_cases = get_sources(
        file_path=FILE_PATH,
        cases_border=CASES_BORDER
    )
